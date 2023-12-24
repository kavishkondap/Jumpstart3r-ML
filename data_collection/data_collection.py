from bs4 import BeautifulSoup
import requests
import lxml
import pandas as pd
import xlrd
from lxml import etree
from openpyxl import Workbook, load_workbook

i = -1
sourceReq = []
soups = []
sources = []
doms = []
successful = []
bioLinks = []
hasTopImage = []
hasTopVideo = []
hasBottomImage = []
hasBottomVideo = []
titles = []
descriptions = []
moneyRaised = []
goals = []
backers = []
updates = []
comments = []
locations = []
bioName = []
bioDescription = []
bioCollaborators = []
bioCollaboratorsLink = []
bioVerifiedIdentity = []
bioLastLogin = []
bioFacebookLink = []
bioFacebookFollowers = []
bioNumBacked = []
bioPastProjectCount = []

# ****** MAKE SURE TO SAVE THE SHEET WITH THE LINKS UNDER THE SAME NAME AS BELOW ("FormattedKickstarter.xlsx") ******

workbook = load_workbook(filename="FormattedKickstarter.xlsx")

numberOfWebsites = 30
startNumber = 0
numNoImage = 0

workbook.read_only
workbook.data_only

sheet = workbook.active

for x in range(startNumber+2, numberOfWebsites+2+startNumber):
    string = "https:" + sheet.cell(x, 1).value
    sources.append(string)
for source in sources:
    sourceReq.append(requests.get(source).text)
for source in sourceReq:
    soups.append(BeautifulSoup(source, 'lxml'))

def findBioInfo (bioLink: BeautifulSoup):
    sourceBio = requests.get(bioLink).text
    bio = BeautifulSoup(sourceBio, 'lxml')
    print ()
    print ("BIO INFO:")
    print()
    name = bio.find ('div', class_='container-flex px3').h1.a.text
    bioName.append(name)
    print ("Name: " + name)
    description = bio.find('div', class_='col col-7 col-post-1 pt3 pb3 pb10-sm').p.text
    bioDescription.append(description)
    print ("Description: " + description)
    try:
        pass
        collab = bio.find('div', class_='col col-7 col-post-1 pt3 pb3 pb10-sm').a.text
        collabLink = bio.find('div', class_='col col-7 col-post-1 pt3 pb3 pb10-sm').a['href']
        collabLink = "https://www.kickstarter.com" + collabLink
        bioCollaborators.append(collab)
        bioCollaboratorsLink.append(collabLink)
        print ('Collaborators: ')
        print (collab)
        print (collabLink)
    except Exception as collab:
        bioCollaborators.append('NO COLLABORATORS')
        bioCollaboratorsLink.append('NO COLLABORATOR LINK')
        print("NO COLLABORATORS")
    except Exception as collabLink:
        bioCollaboratorsLink.append('NO COLLABORATOR LINK')
        print("NO COLLAB LINK")
    except Exception as e:
        print ("JUST FAILED")
    identity = bio.find('div', class_='creator-bio-details col col-4 pt3 pb3 pb10-sm').find('div', class_='verified pb2 border-bottom f5').span.text
    bioVerifiedIdentity.append(identity)
    print ("Verified Identity: " + identity)
    lastLogin = bio.find('div', class_='creator-bio-details col col-4 pt3 pb3 pb10-sm').find('div', class_='last-login py2 border-bottom f5').time.text
    bioLastLogin.append(lastLogin)
    print ("Last Login: " + lastLogin)
    try:
        facebook = bio.find('div', class_='creator-bio-details col col-4 pt3 pb3 pb10-sm').find('div', class_='facebook py2 border-bottom f5').a['href']
        bioFacebookLink.append(facebook)
        print(facebook)
        facebookFollower = bio.find('div', class_='creator-bio-details col col-4 pt3 pb3 pb10-sm').find('div', class_='facebook py2 border-bottom f5').a.text
        bioFacebookFollowers.append (facebookFollower)
        print (facebookFollower)
    except Exception as e:
        bioFacebookLink.append(None)
        bioFacebookFollowers.append(None)
        print("No Facebook Link Present")
    numBacked = bio.find('div', class_='creator-bio-details col col-4 pt3 pb3 pb10-sm').find('div',
                                                                                             class_='created-projects py2 f5 mb3').text
    splitSpaces = numBacked.split(' ')
    try:
        if 'First' in numBacked:
            bioPastProjectCount.append('0')
        else:
            bioPastProjectCount.append(splitSpaces [0])
    except Exception as e:
        bioPastProjectCount.append(None)
        count = 0
    try:
        splitPeriod = numBacked.split('·') [1]
        splitSpacesFormatted = splitPeriod.split(' ')[0]

        print(splitSpacesFormatted)
        bioNumBacked.append(splitSpacesFormatted)
    except Exception as e:
        bioNumBacked.append(None)
        print ('failed num backed')

# ********* STARTING ITERATIONS ********* (NOTHING TO BE CHANGED)
for soup in soups: # ****** FINDING INFO FOR EACH SOURCE ****** (NOTHING TO BE CHANGED)
    i += 1
    dom = etree.HTML(str(soup))
    doms.append(dom)
    hasTopImage.append(False)
    hasTopVideo.append(False)
    hasBottomImage.append(False)
    hasBottomVideo.append(False)
    try:
        success = soup.find('div', class_='Campaign-state-successful')
        if not success == None:
            successful.append(True)
        else:
            successful.append(False)
    except Exception as e:
        successful.append(False)

    # ****** SUCCESSFUL OR NOT ****** (NOTHING TO BE CHANGED)
    print(sources [i]) # ****** URL ****** (NOTHING TO BE CHANGED)
    if (successful[i] == True):
        print("SUCCESSFUL: True")

    else:
        print("SUCCESSFUL: False")

    # ****** TITLE ****** (NOTHING TO BE CHANGED)

    if successful[i]:
        title = soup.find('div', class_='NS_project_profile__title')
        titles.append(title.h2.span.a.text)
        print(title.h2.span.a.text)
    else:
        title = soup.find()
        titles.append(title.h2.text)
        print(title.h2.text)

    # ****** DESCRIPTION ****** (NOTHING TO BE CHANGED)

    if successful[i]:
        descriptionS = soup.find('div', class_='project-profile__blurb editable-field')
        descriptions.append(descriptionS.span.text)
        print(descriptionS.span.text)
    else:
        descriptionF = soup.find('div', class_='grid-container flex flex-column').find('div', class_='grid-col-10 grid-col-10-lg grid-col-offset-1-md block-md order2-md type-center')
        descriptions.append(descriptionF.p.text)
        print(descriptionF.p.text)

    # ****** MONEY RAISED & GOAL ****** (NOTHING TO BE CHANGED)
    hasGoaled = False
    if successful[i]:
        for x in soup.find_all('div', class_='mb3'):
            try:
                pass
                print("Value raised: " + x.h3.span.text)
                moneyRaised.append(x.h3.span.text)
                print ("Goal: " + x.div.span.text)
                hasGoaled = True
                goals.append(x.div.span.text)
                break
            except Exception as e:
                x = None
        if not hasGoaled:
            goals.append (None)
    else:
        divMoney = soup.find ('div', class_='grid-row order2-md hide-lg')
        divMoney = divMoney.find ('div', class_='flex flex-column-lg mb4 mb5-sm')
        divMoney = divMoney.find ('div', class_='mb4-lg')
        divMoney = divMoney.find('div', class_='flex items-center')
        moneyVal = divMoney.find ('span', class_='soft-black')#.text
        print(moneyVal)
        moneyXpath = '//*[@id="react-project-header"]/div/div/div[3]/div/div[2]/div[1]/div[2]/span/span'
        moneyRaised.append(None)
        for x in dom.xpath(moneyXpath):
            print(x.text)
        try:
            goalXpath = '//*[@id="react-project-header"]/div/div/div[3]/div/div[2]/div[1]/span/span[1]/span'
            goal = dom.xpath (goalXpath)[0].text
            goals.append (goal)
            print('Goal:')
            print(goal)
        except Exception as e:
            goals.append(None)

    # ****** NUMBER OF BACKERS ****** (NOTHING TO BE CHANGED)

    if successful[i]:
        backer = soup.find('div', class_='mb0')
        try:
            pass
            print("Backers: " + backer.h3.text)
            backers.append(backer.h3.text)
        except Exception as e:
            print ("FAILED")
    else:
        backersXpath = '/html/body/main/div/div/div[1]/div/div/div[3]/div/div[2]/div[2]/div'
        backersVal = dom.xpath (backersXpath)[0].text
        backers.append(None)
        print("Backers")
        print (backersVal)

    # ****** UPDATES & COMMENTS ****** (NOTHING TO BE CHANGED)

    if successful[i]:
        updatesComments = soup.find('div', class_='campaign-side-nav project-nav__links')
        update = updatesComments.find('a',
                                       class_='js-analytics-section js-load-project-content js-load-project-updates mx3 project-nav__link--updates tabbed-nav__link type-14').text

        formattedUpdate = update [8:update.__len__()-1]
        updates.append(formattedUpdate)
        print(formattedUpdate)
        comment = updatesComments.find('a',
                                        class_='js-analytics-section js-load-project-comments js-load-project-content mx3 project-nav__link--comments tabbed-nav__link type-14').text
        formattedComment = comment [9:comment.__len__()]
        comments.append(formattedComment)
        print(formattedComment)
    else:
        updatesComments = soup.find('div', class_='project-nav__links')
        update = updatesComments.find('a',
                                       class_='js-analytics-section js-load-project-content js-load-project-updates mx3 project-nav__link--updates tabbed-nav__link type-14').text
        formattedUpdate = update[8:update.__len__() - 1]
        updates.append(formattedUpdate)
        print(formattedUpdate)

        comment = updatesComments.find('a',
                                        class_='js-analytics-section js-load-project-comments js-load-project-content mx3 project-nav__link--comments tabbed-nav__link type-14').text
        formattedComment = comment[9:comment.__len__()]
        comments.append(formattedComment)
        print(formattedComment)

    # ****** LOCATION ****** (NOTHING TO BE CHANGED)

    if successful[i]:
        locDiv = soup.find('div', class_='NS_projects__description_section m-auto')
        locA = locDiv.find_all('a', class_='grey-dark mr3 nowrap type-12')
        for x in locA:
            if ',' in x.text:
                locations.append(x.text)
                print("Location: " + x.text)
    else:
        locDiv = soup.find('div',
                           class_='order1-md hide-lg border-top border-bottom border-top-none-md border-none-lg nested-full-width-xs nested-full-width-sm nested-full-width-md mb4 mb5-sm mb0-md')
        locA = locDiv.find_all('a', class_='nowrap navy-700 flex items-center medium mr3 type-12 keyboard-focusable')
        for x in locA:
            if ',' in x.text:
                locations.append(x.text)
                print("Location: " + x.text)

    # ****** IMAGE & VIDEO ****** (NOTHING TO BE CHANGED)

    if (successful [i]):
        try:
            pass
            topImg = soup.find('div', class_='project-profile__content').find('div', class_='grid-container pb3 pb10-sm').find('img', class_='js-feature-image')
            topVid = soup.find('div', class_='project-profile__content').find('div', class_='grid-container pb3 pb10-sm').find('div', class_='video-player')
            if not topVid == None:
                hasTopVideo [i] = True
                hasTopImage [i] = True
                print ("Has Top Video: True")
                print ("Has Top Image: True")
            elif not topImg == None:
                hasTopVideo [i] = False
                hasTopImage [i] = True
                print("Has Top Video: False")
                print("Has Top Image: True")
            else:
                hasTopImage [i] = False
                hasTopVideo [i] = False
                print("Has Top Video: False")
                print("Has Top Image: False")
        except Exception as topVid:
            hasTopVideo [i] = False
            print ("Has Top Video: False")
            if not topImg == None:
                hasTopImage [i] = True
                print ("Has Top Image: True")
            else:
                hasTopImage [i] = False
                print("Has Top Image: False")
        except Exception as topImg:
            hasTopVideo [i] = False
            hasTopImage [i] = False
            print("Has Top Video: False")
            print("Has Top Image: False")
        try:
            pass
            bottomImg = soup.find('div', class_='NS_projects__description_section m-auto').find('div', class_='mb3 aspect-ratio aspect-ratio--16x9').find('img')
            bottomVid = soup.find('div', class_='NS_projects__description_section m-auto').find('div', class_='mb3 aspect-ratio aspect-ratio--16x9').find('div', class_='video-player')
            if not bottomVid == None:
                hasBottomVideo[i] = True
                hasBottomImage[i] = True
                print("Has Bottom Video: True")
                print("Has Bottom Image: True")
            elif not bottomImg == None:
                hasBottomVideo[i] = False
                hasBottomImage[i] = True
                print("Has Bottom Video: False")
                print("Has Bottom Image: True")
            else:
                hasBottomImage[i] = False
                hasBottomVideo[i] = False
                print("Has Bottom Video: False")
                print("Has Bottom Image: False")
        except Exception as e:
            hasBottomImage[i] = False
            hasBottomVideo[i] = False
            print("Has Bottom Video: False")
            print("Has Bottom Image: False")
    else:
        try:
            pass
            topImg = soup.find ('div', class_='bg-grey-100').find('div', class_='grid-row grid-row mb5-lg mb0-md order1-md order2-lg').find('img')
            topVid = soup.find ('div', class_='bg-grey-100').find('div', class_='').find('div', class_='grid-row grid-row mb5-lg mb0-md order1-md order2-lg').find('video')
            if not topVid == None:
                hasTopVideo[i] = True
                hasTopImage[i] = True
                print("Has Top Video: True")
                print("Has Top Image: True")
            elif not topImg == None:
                hasTopVideo[i] = False
                hasTopImage[i] = True
                print("Has Top Video: False")
                print("Has Top Image: True")
            else:
                hasTopImage[i] = False
                hasTopVideo[i] = False
                print("Has Top Video: False")
                print("Has Top Image: False")
        except Exception as topVid:
            hasTopVideo[i] = False
            print("Has Top Video: False")
            if not topImg == None:
                hasTopImage[i] = True
                print("Has Top Image: True")
            else:
                hasTopImage[i] = False
                print("Has Top Image: False")
        except Exception as topImg:
            hasTopVideo[i] = False
            hasTopImage[i] = False
            print("Has Top Video: False")
            print("Has Top Image: False")
        try:
            pass
            bottomImg = soup.find('div', class_='NS_projects__description_section').find('div', class_='supporting-material-callout mb3 mb10-sm').find('div', class_='supporting-material-callout__image').find('img')
            bottomVid = soup.find('div', class_='NS_projects__description_section').find('div', class_='supporting-material-callout mb3 mb10-sm').find('video')
            if not bottomVid == None:
                hasBottomVideo[i] = True
                hasBottomImage[i] = True
                print("Has Bottom Video: True")
                print("Has Bottom Image: True")
            elif not bottomImg == None:
                hasBottomVideo[i] = False
                hasBottomImage[i] = True
                print("Has Bottom Video: False")
                print("Has Bottom Image: True")
            else:
                hasBottomImage[i] = False
                hasBottomVideo[i] = False
                print("Has Bottom Video: False")
                print("Has Bottom Image: False")
        except Exception as e:
            hasBottomImage[i] = False
            hasBottomVideo[i] = False
            print("Has Bottom Video: False")
            print("Has Bottom Image: False")


    # ****** PERSON INFO ****** (NOTHING TO BE CHANGED)

    if (successful[i]):
        link = soup.find('div', class_='NS_projects__creator_spotlight mobile-center').find('div', class_='mobile-hide').a['href']
        link = "https://www.kickstarter.com" + link
        print(link)
        bioLinks.append(link)
        findBioInfo (link)
    else:
        bioLinks.append (None)
        bioName.append (None)
        bioDescription.append (None)
        bioCollaborators.append (None)
        bioCollaboratorsLink.append(None)
        bioVerifiedIdentity.append (None)
        bioLastLogin.append (None)
        bioFacebookLink.append (None)
        bioFacebookFollowers.append(None)
        bioNumBacked.append (None)
        bioPastProjectCount.append(None)

    print()
    print()
    print()
    print()

data = {'Source': sources,
        'Successful': successful,
        'Title': titles,
        'Descriptions': descriptions,
        'Money Raised': moneyRaised,
        'Goal': goals,
        'Backers': backers,
        'Updates': updates,
        'Comments': comments,
        'Top Video': hasTopVideo,
        'Top Image': hasTopImage,
        'Bottom Video': hasBottomVideo,
        'Bottom Image': hasBottomImage,
        'Location': locations,
        'Bio Link': bioLinks,
        'Bio Name': bioName,
        'Bio Description': bioDescription,
        'Bio Collaborators': bioCollaborators,
        'Bio Collaborators Link': bioCollaboratorsLink,
        'Bio Verified Identity': bioVerifiedIdentity,
        'Bio Last Login': bioLastLogin,
        'Bio Facebook Link': bioFacebookLink,
        'Bio Facebook Followers': bioFacebookFollowers,
        'Bio Number Backed': bioNumBacked,
        'Bio Past Project Count': bioPastProjectCount
       }

df = pd.DataFrame(data, columns = ['Source',
        'Successful',
        'Title',
        'Descriptions',
        'Money Raised',
        'Goal',
        'Backers',
        'Updates',
        'Comments',
        'Top Video',
        'Top Image',
        'Bottom Image',
        'Bottom Video',
        'Location',
        'Bio Link',
        'Bio Name',
        'Bio Description',
        'Bio Collaborators',
        'Bio Collaborators Link',
        'Bio Verified Identity',
        'Bio Last Login',
        'Bio Facebook Link',
        'Bio Facebook Followers',
        'Bio Number Backed',
        'Bio Past Project Count'])


# ******** EXPORTING TO EXCEL ******** PLEASE CHANGE THE ADDRESS OF THE SAVED FILE TO A LOCATION ON YOUR COMPUTER ********

df.to_excel (r'C:/Users/kavis/Desktop/KAVISH/General Stuff/WebScraping/webscrape_data.xlsx', index = False, header=True)

# ******* THE LOCATION TEXT SHOULD BE IN THE FORM r'(FILE ADDRESS)/(FILE NAME)' ******

l = []
l.append(len(sourceReq))
l.append(len(soups))
l.append(len(sources))
l.append(len(doms))
l.append(len(successful))
l.append(len(bioLinks))
l.append(len(hasTopImage))
l.append(len(hasTopVideo))
l.append(len(hasBottomImage))
l.append(len(hasBottomVideo))
l.append(len(titles))
l.append(len(descriptions))
l.append(len(moneyRaised))
l.append(len(goals))
l.append(len(backers))
l.append(len(updates))
l.append(len(comments))
l.append(len(locations))
l.append(len(bioName))
l.append(len(bioDescription))
l.append(len(bioCollaborators))
l.append(len(bioCollaboratorsLink))
l.append(len(bioVerifiedIdentity))
l.append(len(bioLastLogin))
l.append(len(bioFacebookLink))
l.append(len(bioFacebookFollowers))
l.append(len(bioNumBacked))
l.append(len(bioPastProjectCount))

print (l)